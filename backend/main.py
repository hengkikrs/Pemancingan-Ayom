"""
Pemancingan Ayom — FastAPI Backend
Handles: auth, report generation (Excel + PDF)
Run: uvicorn main:app --reload --port 8000
"""

from fastapi import FastAPI, HTTPException, Depends, Query, Request
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime, timedelta
import io, os, hashlib, hmac, base64, json, traceback

# Load .env file otomatis
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("✅ .env loaded")
except ImportError:
    print("⚠️  python-dotenv tidak terinstall, jalankan: pip install python-dotenv")

# ── Optional imports (install separately) ────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    from supabase import create_client, Client
    SUPABASE_OK = True
except ImportError:
    SUPABASE_OK = False

# ── Supabase client ───────────────────────────────────────
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")  # service role key for backend

def get_supabase() -> "Client":
    if not SUPABASE_OK:
        raise HTTPException(500, "supabase-py not installed")
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise HTTPException(500, "Supabase env vars not set")
    return create_client(SUPABASE_URL, SUPABASE_KEY)

# ── App ───────────────────────────────────────────────────
app = FastAPI(title="Pemancingan Ayom API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Global exception handler: tampilkan traceback di log ──
from fastapi import Request as _Request
from fastapi.responses import JSONResponse as _JSONResponse
import traceback as _tb

@app.exception_handler(Exception)
async def _global_exc(request: _Request, exc: Exception):
    trace = _tb.format_exc()
    print(f"\n{'='*60}\nERROR: {request.url}\n{trace}\n{'='*60}")
    return _JSONResponse(status_code=500, content={"detail": str(exc), "trace": trace})

# ── Helpers ───────────────────────────────────────────────
def fmt_idr(n):
    if n is None: return "Rp 0"
    try:
        return f"Rp {int(n):,}".replace(",", ".")
    except:
        return f"Rp {n}"

def calc_galatama(participants: int):
    total   = participants * 10000
    profit  = total * 0.5
    pool    = total * 0.5
    if participants <= 10:
        return dict(total=total, profit=profit, pool=pool,
                    j1=pool*0.6, j2=pool*0.4, j3=None,
                    j1_pct="60%", j2_pct="40%", j3_pct=None)
    return dict(total=total, profit=profit, pool=pool,
                j1=pool*0.5, j2=pool*0.3, j3=pool*0.2,
                j1_pct="50%", j2_pct="30%", j3_pct="20%")

# ── Data fetchers ─────────────────────────────────────────
def fetch_galatama(sb, start: date, end: date):
    res = sb.table("galatama_sessions")\
        .select("*, users(full_name)")\
        .gte("session_date", str(start))\
        .lte("session_date", str(end))\
        .order("session_date").order("session_num")\
        .execute()
    return res.data or []

def fetch_galatama_kasbon(sb, start: date, end: date):
    res = sb.table("galatama_kasbon")\
        .select("*, galatama_sessions(session_date, session_num), users(full_name)")\
        .gte("created_at", str(start))\
        .lte("created_at", str(end) + " 23:59:59")\
        .execute()
    return res.data or []

def fetch_warung(sb, start: date, end: date):
    res = sb.table("warung_transactions")\
        .select("*, users(full_name)")\
        .gte("trans_date", str(start))\
        .lte("trans_date", str(end))\
        .order("trans_date").order("created_at")\
        .execute()
    return res.data or []

def fetch_open_bills(sb):
    res = sb.table("open_bills").select("*").eq("status", "open").execute()
    return res.data or []

# ─────────────────────────────────────────────────────────
# EXCEL REPORT
# ─────────────────────────────────────────────────────────
BLUE_DARK  = "1E3A5F"
BLUE_MID   = "2563EB"
BLUE_LIGHT = "DBEAFE"
BLUE_PALE  = "EFF6FF"
WHITE      = "FFFFFF"
EMERALD    = "059669"
AMBER      = "D97706"
ROSE       = "DC2626"
GRAY       = "6B7280"

def style_header_cell(cell, bg=BLUE_DARK, fg=WHITE, bold=True, size=11):
    cell.font            = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill            = PatternFill("solid", fgColor=bg)
    cell.alignment       = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_cell(cell, bold=False, color="000000", bg=None, align="left"):
    cell.font      = Font(bold=bold, color=color, size=10, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        cell.fill  = PatternFill("solid", fgColor=bg)

def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, row_start, row_end, col_start, col_end):
    for row in ws.iter_rows(min_row=row_start, max_row=row_end,
                             min_col=col_start, max_col=col_end):
        for cell in row:
            cell.border = thin_border()

def add_title_block(ws, title, subtitle, date_range):
    # Row 1: logo area
    ws.merge_cells("A1:H2")
    cell = ws["A1"]
    cell.value     = "🎣  PEMANCINGAN AYOM"
    cell.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=BLUE_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value     = title
    c.font      = Font(bold=True, size=13, color=BLUE_DARK, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=BLUE_PALE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value     = f"{subtitle}  |  {date_range}"
    c.font      = Font(size=10, color=GRAY, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=BLUE_PALE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 18
    return 5  # next row

def generate_excel(start: date, end: date, galatama_rows, warung_rows, kasbon_rows) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: RINGKASAN ────────────────────────────────
    ws = wb.active
    ws.title = "Ringkasan"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    ws.merge_cells("A1:C2")
    c = ws["A1"]
    c.value     = "🎣  PEMANCINGAN AYOM"
    c.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A3:C3")
    c = ws["A3"]
    date_str = f"{start.strftime('%d %b %Y')} — {end.strftime('%d %b %Y')}"
    c.value     = f"LAPORAN RINGKASAN  |  {date_str}"
    c.font      = Font(bold=True, size=11, color=BLUE_DARK, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=BLUE_PALE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Galatama summary
    gal_total    = sum(r.get("total_revenue", r["participants"]*10000) for r in galatama_rows)
    gal_profit   = gal_total * 0.5
    gal_prize    = gal_total * 0.5
    gal_sessions = len(galatama_rows)
    gal_pax      = sum(r["participants"] for r in galatama_rows)

    # Warung summary
    war_rev    = sum(float(r.get("revenue", 0)) for r in warung_rows)
    war_profit = sum(float(r.get("profit", 0)) for r in warung_rows)
    war_cogs   = sum(float(r.get("cogs", 0)) for r in warung_rows)

    total_profit = gal_profit + war_profit

    row = 5
    sections = [
        ("KOLAM GALATAMA", [
            ("Jumlah Sesi", str(gal_sessions)),
            ("Total Pemancing", str(gal_pax)),
            ("Total Omzet", fmt_idr(gal_total)),
            ("Profit Kolam (50%)", fmt_idr(gal_profit)),
            ("Total Hadiah (50%)", fmt_idr(gal_prize)),
        ]),
        ("WARUNG", [
            ("Total Transaksi", str(len(warung_rows))),
            ("Total Omzet", fmt_idr(war_rev)),
            ("HPP", fmt_idr(war_cogs)),
            ("Laba Kotor", fmt_idr(war_profit)),
            ("Margin", f"{round(war_profit/war_rev*100) if war_rev else 0}%"),
        ]),
        ("TOTAL GABUNGAN", [
            ("Total Omzet", fmt_idr(gal_total + war_rev)),
            ("Total Laba Bersih", fmt_idr(total_profit)),
            ("Kasbon Warung Aktif", str(len([r for r in warung_rows if not r.get("is_settled")]))),
        ]),
    ]

    for section_title, items in sections:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value     = section_title
        c.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=BLUE_MID)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20
        row += 1

        for label, value in items:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            style_data_cell(ws[f"A{row}"], bg=BLUE_PALE if row%2==0 else WHITE)
            style_data_cell(ws[f"B{row}"], bold=True, color=BLUE_DARK, align="right", bg=BLUE_PALE if row%2==0 else WHITE)
            apply_border(ws, row, row, 1, 3)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── Sheet 2: GALATAMA ─────────────────────────────────
    wg = wb.create_sheet("Galatama")
    wg.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [12, 8, 10, 14, 14, 14, 14, 14]):
        wg.column_dimensions[get_column_letter(ord(col)-64)].width = w

    nr = add_title_block(wg, "LAPORAN GALATAMA",
                         f"{gal_sessions} sesi | {gal_pax} pemancing",
                         f"{start.strftime('%d/%m/%Y')} – {end.strftime('%d/%m/%Y')}")

    headers = ["Tanggal","Sesi","Peserta","Omzet","Profit Kolam","Pool Hadiah","Juara 1","Juara 2"]
    for i, h in enumerate(headers, 1):
        c = wg.cell(nr, i, h)
        style_header_cell(c, BLUE_MID)
    wg.row_dimensions[nr].height = 22
    nr += 1

    for idx, r in enumerate(galatama_rows):
        g     = calc_galatama(r["participants"])
        bg    = BLUE_PALE if idx % 2 == 0 else WHITE
        vals  = [
            r.get("session_date",""),
            f"Sesi {r['session_num']}",
            r["participants"],
            fmt_idr(g["total"]),
            fmt_idr(g["profit"]),
            fmt_idr(g["pool"]),
            f"{r.get('winner1_name','—')} ({fmt_idr(g['j1'])})",
            f"{r.get('winner2_name','—')} ({fmt_idr(g['j2'])})",
        ]
        for i, v in enumerate(vals, 1):
            c = wg.cell(nr, i, v)
            style_data_cell(c, bg=bg, align="center" if i in (1,2,3) else "left")
        apply_border(wg, nr, nr, 1, 8)
        wg.row_dimensions[nr].height = 18
        nr += 1

    # Totals row
    nr += 1
    wg.cell(nr, 1, "TOTAL").font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    wg.cell(nr, 1).fill = PatternFill("solid", fgColor=BLUE_DARK)
    wg.cell(nr, 1).alignment = Alignment(horizontal="center")
    wg.cell(nr, 3, gal_pax)
    wg.cell(nr, 4, fmt_idr(gal_total))
    wg.cell(nr, 5, fmt_idr(gal_profit))
    wg.cell(nr, 6, fmt_idr(gal_prize))
    for i in range(1, 9):
        c = wg.cell(nr, i)
        if i > 1:
            style_data_cell(c, bold=True, color=WHITE, bg=BLUE_DARK, align="center")
        apply_border(wg, nr, nr, i, i)
    wg.row_dimensions[nr].height = 22

    # ── Sheet 3: WARUNG ───────────────────────────────────
    ww = wb.create_sheet("Warung")
    ww.sheet_view.showGridLines = False
    for col, w in zip(range(1,9), [12,20,12,10,14,14,14,12]):
        ww.column_dimensions[get_column_letter(col)].width = w

    war_total_tx = len(warung_rows)
    nr2 = add_title_block(ww, "LAPORAN WARUNG",
                          f"{war_total_tx} transaksi",
                          f"{start.strftime('%d/%m/%Y')} – {end.strftime('%d/%m/%Y')}")

    h2 = ["Tanggal","Produk","Kategori","Qty","Harga Jual","HPP","Omzet","Laba"]
    for i, h in enumerate(h2, 1):
        c = ww.cell(nr2, i, h)
        style_header_cell(c, BLUE_MID)
    ww.row_dimensions[nr2].height = 22
    nr2 += 1

    for idx, r in enumerate(warung_rows):
        bg   = BLUE_PALE if idx % 2 == 0 else WHITE
        rev  = float(r.get("revenue", 0))
        prof = float(r.get("profit", 0))
        vals = [
            r.get("trans_date",""),
            r.get("product_name",""),
            r.get("category",""),
            f"{r['qty']} {r.get('unit','pcs')}",
            fmt_idr(r.get("harga_jual",0)),
            fmt_idr(r.get("cogs",0)),
            fmt_idr(rev),
            fmt_idr(prof),
        ]
        for i, v in enumerate(vals, 1):
            c = ww.cell(nr2, i, v)
            style_data_cell(c, bg=bg, color=EMERALD if i==8 else "000000",
                            bold=(i==8), align="right" if i >= 5 else "left")
        apply_border(ww, nr2, nr2, 1, 8)
        ww.row_dimensions[nr2].height = 17
        nr2 += 1

    nr2 += 1
    for i, (lbl, val) in enumerate(
        [("",""), ("",""), ("",""), ("",""), ("",""), ("",""),
         ("TOTAL OMZET", fmt_idr(war_rev)), ("TOTAL LABA", fmt_idr(war_profit))], 1):
        c = ww.cell(nr2, i)
        if lbl:
            c.value = lbl if i == 7 else val
            style_data_cell(c, bold=True, color=WHITE, bg=BLUE_DARK if i==7 else EMERALD, align="right")
        apply_border(ww, nr2, nr2, i, i)
    ww.row_dimensions[nr2].height = 22

    # ── Sheet 4: KASBON ───────────────────────────────────
    wk = wb.create_sheet("Kasbon")
    wk.sheet_view.showGridLines = False
    for col, w in zip(range(1,7), [12,20,14,12,20,20]):
        wk.column_dimensions[get_column_letter(col)].width = w

    nr3 = add_title_block(wk, "LAPORAN KASBON", "Galatama + Warung",
                          f"{start.strftime('%d/%m/%Y')} – {end.strftime('%d/%m/%Y')}")

    h3 = ["Tanggal","Nama","Jumlah","Status","Catatan","Oleh"]
    for i, h in enumerate(h3, 1):
        c = wk.cell(nr3, i, h)
        style_header_cell(c, BLUE_MID)
    wk.row_dimensions[nr3].height = 22
    nr3 += 1

    warung_kasbon = [r for r in warung_rows if r.get("payment_type") == "kasbon"]
    all_kasbon = kasbon_rows + warung_kasbon
    for idx, r in enumerate(all_kasbon):
        bg      = BLUE_PALE if idx % 2 == 0 else WHITE
        status  = r.get("status","open")
        is_settled = status == "settled" or r.get("is_settled", False)
        vals = [
            r.get("trans_date", r.get("created_at",""))[:10],
            r.get("kasbon_name", r.get("angler_name","")),
            fmt_idr(float(r.get("amount", r.get("revenue", 0)))),
            "✅ Lunas" if is_settled else "⏳ Belum Lunas",
            r.get("notes","") or "-",
            r.get("created_by",""),
        ]
        for i, v in enumerate(vals, 1):
            c = wk.cell(nr3, i, v)
            col_color = EMERALD if (i==4 and is_settled) else AMBER if i==4 else "000000"
            style_data_cell(c, bg=bg, color=col_color, bold=(i==4))
        apply_border(wk, nr3, nr3, 1, 6)
        wk.row_dimensions[nr3].height = 17
        nr3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────
# PDF REPORT
# ─────────────────────────────────────────────────────────
def generate_pdf(start: date, end: date, galatama_rows, warung_rows, kasbon_rows) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.8*cm, rightMargin=1.8*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    BLUE_RL  = colors.HexColor("#1E3A5F")
    BLUE2_RL = colors.HexColor("#2563EB")
    PALE_RL  = colors.HexColor("#EFF6FF")
    EMR_RL   = colors.HexColor("#059669")
    AMB_RL   = colors.HexColor("#D97706")
    GRAY_RL  = colors.HexColor("#6B7280")

    h1_style = ParagraphStyle("H1", fontSize=18, fontName="Helvetica-Bold",
                               textColor=colors.white, alignment=TA_CENTER, spaceAfter=2)
    h2_style = ParagraphStyle("H2", fontSize=12, fontName="Helvetica-Bold",
                               textColor=BLUE_RL, spaceAfter=6, spaceBefore=12)
    body_style = ParagraphStyle("Body", fontSize=9, fontName="Helvetica",
                                 textColor=colors.black, spaceAfter=2)
    sub_style  = ParagraphStyle("Sub", fontSize=9, fontName="Helvetica",
                                 textColor=GRAY_RL, alignment=TA_CENTER)

    def tbl_style(header_color=BLUE2_RL):
        return TableStyle([
            ("BACKGROUND", (0,0), (-1,0), header_color),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,0), 9),
            ("ALIGN",      (0,0), (-1,0), "CENTER"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, PALE_RL]),
            ("FONTNAME",   (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",   (0,1), (-1,-1), 8.5),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.HexColor("#D1D5DB")),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("RIGHTPADDING",(0,0), (-1,-1), 6),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ])

    story = []
    date_range = f"{start.strftime('%d %b %Y')} — {end.strftime('%d %b %Y')}"

    # ── Cover block ───────────────────────────────────────
    cover_data = [
        [Paragraph("🎣  PEMANCINGAN AYOM", h1_style)],
        [Paragraph(f"Laporan Keuangan  |  {date_range}", sub_style)],
        [Paragraph(f"Digenerate: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_style)],
    ]
    cover_tbl = Table(cover_data, colWidths=[doc.width])
    cover_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), BLUE_RL),
        ("BACKGROUND", (0,1), (-1,-1), PALE_RL),
        ("TOPPADDING", (0,0), (0,0), 14),
        ("BOTTOMPADDING", (0,0), (0,0), 14),
        ("TOPPADDING", (0,1), (-1,-1), 6),
        ("BOTTOMPADDING", (0,-1), (-1,-1), 8),
        ("BOX", (0,0), (-1,-1), 1, BLUE_RL),
    ]))
    story.append(cover_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Ringkasan KPI ─────────────────────────────────────
    gal_total  = sum(r["participants"]*10000 for r in galatama_rows)
    gal_profit = gal_total * 0.5
    war_rev    = sum(float(r.get("revenue",0)) for r in warung_rows)
    war_profit = sum(float(r.get("profit",0)) for r in warung_rows)

    story.append(Paragraph("RINGKASAN EKSEKUTIF", h2_style))
    kpi_data = [
        ["Metrik", "Galatama", "Warung", "Total"],
        ["Omzet",      fmt_idr(gal_total),  fmt_idr(war_rev),    fmt_idr(gal_total+war_rev)],
        ["Laba Bersih",fmt_idr(gal_profit), fmt_idr(war_profit), fmt_idr(gal_profit+war_profit)],
        ["Sesi/Transaksi", str(len(galatama_rows)), str(len(warung_rows)), "—"],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[4*cm, 3.8*cm, 3.8*cm, 3.8*cm])
    kpi_tbl.setStyle(tbl_style())
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Galatama Detail ───────────────────────────────────
    if galatama_rows:
        story.append(HRFlowable(width="100%", thickness=1, color=PALE_RL))
        story.append(Paragraph("KOLAM GALATAMA", h2_style))
        g_data = [["Tanggal","Sesi","Peserta","Omzet","Profit","Juara 1","Juara 2"]]
        for r in galatama_rows:
            g = calc_galatama(r["participants"])
            g_data.append([
                str(r.get("session_date","")),
                f"S{r['session_num']}",
                str(r["participants"]),
                fmt_idr(g["total"]),
                fmt_idr(g["profit"]),
                f"{r.get('winner1_name','—')} ({fmt_idr(g['j1'])})",
                f"{r.get('winner2_name','—')} ({fmt_idr(g['j2'])})",
            ])
        # totals
        g_data.append(["TOTAL","",str(sum(r["participants"] for r in galatama_rows)),
                        fmt_idr(gal_total), fmt_idr(gal_profit),"",""])

        g_tbl = Table(g_data, colWidths=[2.2*cm,1.2*cm,1.5*cm,2.8*cm,2.8*cm,3.5*cm,3.5*cm])
        st = tbl_style()
        st.add("BACKGROUND", (0,-1), (-1,-1), BLUE_RL)
        st.add("TEXTCOLOR",  (0,-1), (-1,-1), colors.white)
        st.add("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold")
        g_tbl.setStyle(st)
        story.append(g_tbl)
        story.append(Spacer(1, 0.5*cm))

    # ── Warung Detail ─────────────────────────────────────
    if warung_rows:
        story.append(HRFlowable(width="100%", thickness=1, color=PALE_RL))
        story.append(Paragraph("WARUNG", h2_style))
        w_data = [["Tanggal","Produk","Kategori","Qty","Omzet","Laba"]]
        for r in warung_rows:
            w_data.append([
                str(r.get("trans_date","")),
                r.get("product_name",""),
                r.get("category",""),
                f"{r['qty']} {r.get('unit','pcs')}",
                fmt_idr(float(r.get("revenue",0))),
                fmt_idr(float(r.get("profit",0))),
            ])
        w_data.append(["TOTAL","","","",fmt_idr(war_rev),fmt_idr(war_profit)])
        w_tbl = Table(w_data, colWidths=[2.2*cm,4.5*cm,2.2*cm,2*cm,3.2*cm,3.2*cm])
        st2 = tbl_style()
        st2.add("BACKGROUND", (0,-1), (-1,-1), BLUE_RL)
        st2.add("TEXTCOLOR",  (0,-1), (-1,-1), colors.white)
        st2.add("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold")
        st2.add("TEXTCOLOR",  (5,1), (5,-2), EMR_RL)
        w_tbl.setStyle(st2)
        story.append(w_tbl)
        story.append(Spacer(1, 0.5*cm))

    # ── Kasbon ────────────────────────────────────────────
    all_kasbon = kasbon_rows + [r for r in warung_rows if r.get("payment_type")=="kasbon"]
    if all_kasbon:
        story.append(HRFlowable(width="100%", thickness=1, color=PALE_RL))
        story.append(Paragraph("KASBON", h2_style))
        k_data = [["Tanggal","Nama","Jumlah","Status","Catatan"]]
        for r in all_kasbon:
            settled = r.get("status")=="settled" or r.get("is_settled",False)
            k_data.append([
                str(r.get("trans_date", r.get("created_at",""))[:10]),
                r.get("kasbon_name", r.get("angler_name","")),
                fmt_idr(float(r.get("amount", r.get("revenue",0)))),
                "Lunas" if settled else "Belum Lunas",
                r.get("notes","") or "-",
            ])
        k_tbl = Table(k_data, colWidths=[2.2*cm,4*cm,3*cm,2.5*cm,5.8*cm])
        st3 = tbl_style()
        k_tbl.setStyle(st3)
        story.append(k_tbl)

    # ── Footer ────────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    footer_data = [[Paragraph(
        f"Dokumen ini digenerate otomatis oleh sistem Pemancingan Ayom pada "
        f"{datetime.now().strftime('%d %B %Y pukul %H:%M WIB')}",
        ParagraphStyle("footer", fontSize=8, fontName="Helvetica",
                       textColor=GRAY_RL, alignment=TA_CENTER)
    )]]
    f_tbl = Table(footer_data, colWidths=[doc.width])
    f_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), PALE_RL),
        ("TOPPADDING",(0,0),(-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ("BOX",(0,0),(-1,-1),0.5, BLUE2_RL),
    ]))
    story.append(f_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────
# API ENDPOINTS
# ─────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "excel": EXCEL_OK, "pdf": PDF_OK, "supabase": SUPABASE_OK}

@app.get("/api/report/excel")
def report_excel(
    start: date = Query(default=date.today()),
    end:   date = Query(default=date.today()),
    period: str = Query(default="daily"),  # daily | weekly
):
    if not EXCEL_OK:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    if period == "weekly":
        end = start + timedelta(days=6)

    sb = get_supabase()
    galatama = fetch_galatama(sb, start, end)
    warung   = fetch_warung(sb, start, end)
    kasbon   = fetch_galatama_kasbon(sb, start, end)

    data = generate_excel(start, end, galatama, warung, kasbon)
    filename = f"laporan_ayom_{start}_{end}.xlsx"

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.get("/api/report/pdf")
def report_pdf(
    start: date = Query(default=date.today()),
    end:   date = Query(default=date.today()),
    period: str = Query(default="daily"),
):
    if not PDF_OK:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    if period == "weekly":
        end = start + timedelta(days=6)

    sb = get_supabase()
    galatama = fetch_galatama(sb, start, end)
    warung   = fetch_warung(sb, start, end)
    kasbon   = fetch_galatama_kasbon(sb, start, end)

    data = generate_pdf(start, end, galatama, warung, kasbon)
    filename = f"laporan_ayom_{start}_{end}.pdf"

    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ── Auth verify endpoint (dibutuhkan frontend) ────────────
from pydantic import BaseModel as BM

class LoginReq(BM):
    username: str
    password: str

# Demo passwords (fallback jika Supabase users table belum diisi)
DEMO_PASSWORDS = {
    "inay": "ayom2024",
    "wembi": "wembi",
    "gun": "gun",
}

@app.post("/api/auth/verify")
def auth_verify(req: LoginReq):
    """Verifikasi password. Coba bcrypt dulu, fallback ke demo passwords."""
    # Coba bcrypt dari Supabase
    if SUPABASE_OK and SUPABASE_URL and SUPABASE_KEY:
        try:
            sb = create_client(SUPABASE_URL, SUPABASE_KEY)
            res = sb.table("users").select("password_hash").eq("username", req.username).limit(1).execute()
            if res.data:
                pw_hash = res.data[0]["password_hash"]
                try:
                    import bcrypt
                    valid = bcrypt.checkpw(req.password.encode(), pw_hash.encode())
                    return {"valid": valid}
                except ImportError:
                    pass  # bcrypt tidak tersedia, lanjut ke fallback
        except Exception:
            pass  # Supabase error, lanjut ke fallback

    # Fallback: cek demo passwords
    valid = DEMO_PASSWORDS.get(req.username) == req.password
    return {"valid": valid}